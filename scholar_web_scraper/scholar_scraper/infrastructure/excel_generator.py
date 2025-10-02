"""
Gerador de planilhas Excel com dados de artigos analisados.

Implementa criação de planilhas formatadas com dados dinâmicos
baseados nas configurações de colunas do usuário.
"""

import logging
from typing import List, Dict, Any, Optional
from datetime import datetime, date
import tempfile
import os

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from scholar_scraper.domain.entities import Article, ColumnConfig
from scholar_scraper.domain.value_objects import AnalysisResult


logger = logging.getLogger(__name__)


class ExcelGenerator:
    """
    Gerador de planilhas Excel para dados de artigos científicos.
    
    Cria planilhas formatadas com colunas dinâmicas baseadas
    nas configurações do usuário.
    """
    
    def __init__(self):
        """Inicializa o gerador Excel."""
        # Estilos para formatação
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
    def generate_excel(self,
                      articles: List[Article],
                      analysis_results: List[AnalysisResult],
                      columns: List[ColumnConfig],
                      search_query: str,
                      filter_criteria: str,
                      format_type: str = 'xlsx') -> str:
        """
        Gera planilha Excel ou CSV com artigos analisados.
        
        Args:
            articles: Lista de artigos processados
            analysis_results: Resultados da análise por IA
            columns: Configuração das colunas
            search_query: Texto da busca realizada
            filter_criteria: Critérios de filtro aplicados
            format_type: Formato do arquivo ('xlsx' ou 'csv')
            
        Returns:
            Caminho do arquivo gerado
        """
        try:
            # Constrói DataFrame com dados
            df = self._build_dataframe(articles, analysis_results, columns)
            
            # Determina extensão e prefixo
            extension = '.csv' if format_type == 'csv' else '.xlsx'
            
            # Cria arquivo temporário
            temp_file = tempfile.NamedTemporaryFile(
                delete=False,
                suffix=extension,
                prefix='scholar_scraper_'
            )
            temp_file.close()
            
            if format_type == 'csv':
                # Limpa dados para evitar problemas com separador
                cleaned_df = self._clean_dataframe_for_csv(df)
                # Gera CSV com ponto e vírgula como separador
                cleaned_df.to_csv(temp_file.name, index=False, encoding='utf-8', sep=';')
                logger.info(f"CSV gerado: {temp_file.name}")
            else:
                # Gera Excel com múltiplas abas e formatação
                with pd.ExcelWriter(temp_file.name, engine='openpyxl') as writer:
                    # Aba principal com dados
                    df.to_excel(writer, sheet_name='Artigos Analisados', index=False)
                    
                    # Aba com metadados
                    metadata_df = self._build_metadata_dataframe(
                        search_query, filter_criteria, len(articles), columns
                    )
                    metadata_df.to_excel(writer, sheet_name='Metadados', index=False)
                    
                    # Aplica formatação
                    self._format_workbook(writer.book, df, columns)
                
                logger.info(f"Arquivo gerado: {temp_file.name}")
                
            return temp_file.name
            
        except Exception as e:
            logger.error(f"Erro ao gerar Excel: {e}")
            raise
            
    def _build_dataframe(self, 
                        articles: List[Article],
                        analysis_results: List[AnalysisResult],
                        columns: List[ColumnConfig]) -> pd.DataFrame:
        """
        Constrói DataFrame com dados dos artigos.
        
        Args:
            articles: Lista de artigos
            analysis_results: Resultados das análises
            columns: Configurações de colunas
            
        Returns:
            DataFrame formatado
        """
        rows = []
        
        for i, article in enumerate(articles):
            row = {}
            
            # Colunas básicas do artigo (dados do scraping)
            metadata = article.metadata if hasattr(article, 'metadata') else None
            
            # Título - limpo e sem caracteres especiais
            title = self._clean_text(getattr(metadata, 'title', '') if metadata else '')
            row['Título'] = title if title else 'Título não disponível'
            
            # URL - limpa e validada
            url = self._clean_url(getattr(article, 'scholar_url', '') if hasattr(article, 'scholar_url') else '')
            row['URL'] = url
            
            # Autores - tratamento especial para evitar problemas de formatação
            if metadata and metadata.authors:
                authors_clean = [self._clean_text(author) for author in metadata.authors if author]
                row['Autores'] = ', '.join(authors_clean[:3])  # Máximo 3 autores para evitar linhas muito longas
            else:
                row['Autores'] = 'Autores não especificados'
            
            # Ano de publicação
            row['Ano de Publicação'] = str(metadata.year) if metadata and metadata.year else 'Não especificado'
            
            # Resultado da análise IA
            if i < len(analysis_results):
                result = analysis_results[i]
                
                # Confiança da IA
                row['Confiança da IA'] = f"{result.confidence_score:.2f}"
                
                # Busca valores específicos da análise IA
                row['Descrição'] = self._clean_text(result.get_column_value('Descrição') or 
                                                  result.get_column_value('Resumo') or 
                                                  getattr(metadata, 'snippet', '') if metadata else '')
                
                row['Tecnologias Usadas'] = self._clean_text(result.get_column_value('Tecnologia Principal') or 
                                                           result.get_column_value('Tecnologias Usadas') or 'Não especificado')
                
                row['Tem Aplicativo Móvel?'] = self._format_boolean_value(result.get_column_value('Tem Aplicativo Móvel?') or 
                                                                        result.get_column_value('Aplicativo Móvel'))
                
                row['Tem Painel de Gestão?'] = self._format_boolean_value(result.get_column_value('Tem Painel de Gestão?') or 
                                                                        result.get_column_value('Painel Gestão'))
                
                row['Atende ao Filtro?'] = 'Sim' if result.meets_filter else 'Não'
                
            else:
                # Valores padrão quando análise IA não disponível
                row['Confiança da IA'] = '0.00'
                row['Descrição'] = self._clean_text(getattr(metadata, 'snippet', '') if metadata else '') or 'Descrição não disponível'
                row['Tecnologias Usadas'] = 'Não analisado'
                row['Tem Aplicativo Móvel?'] = 'Não analisado'
                row['Tem Painel de Gestão?'] = 'Não analisado'
                row['Atende ao Filtro?'] = 'Não analisado'
                    
            rows.append(row)
            
        return pd.DataFrame(rows)
        
    def _clean_text(self, text: str) -> str:
        """
        Limpa texto removendo caracteres problemáticos para CSV.
        
        Args:
            text: Texto a ser limpo
            
        Returns:
            Texto limpo
        """
        if not text or text == 'None':
            return ''
            
        text = str(text).strip()
        
        # Remove caracteres problemáticos para CSV
        text = text.replace(';', ',')  # Substitui ponto e vírgula por vírgula
        text = text.replace('\n', ' ')  # Remove quebras de linha
        text = text.replace('\r', ' ')  # Remove carriage returns
        text = text.replace('\t', ' ')  # Remove tabs
        text = text.replace('"', "'")   # Substitui aspas duplas por simples
        
        # Remove espaços duplos
        while '  ' in text:
            text = text.replace('  ', ' ')
            
        return text
        
    def _clean_url(self, url: str) -> str:
        """
        Limpa e valida URL.
        
        Args:
            url: URL a ser limpa
            
        Returns:
            URL limpa ou string vazia se inválida
        """
        if not url or url == 'None':
            return ''
            
        url = str(url).strip()
        
        # Verifica se é uma URL válida
        if not (url.startswith('http://') or url.startswith('https://')):
            return ''
            
        # Remove caracteres problemáticos
        url = url.replace(';', '')
        url = url.replace('\n', '')
        url = url.replace('\r', '')
        url = url.replace(' ', '')
        
        return url
        
    def _format_boolean_value(self, value: Any) -> str:
        """
        Formata valores booleanos de forma consistente.
        
        Args:
            value: Valor a ser formatado
            
        Returns:
            'Sim', 'Não' ou 'Não especificado'
        """
        if value is None or value == '' or value == 'None':
            return 'Não especificado'
            
        if isinstance(value, bool):
            return 'Sim' if value else 'Não'
            
        if isinstance(value, str):
            value_lower = value.lower().strip()
            if value_lower in ['true', 'sim', 'yes', '1', 'verdadeiro']:
                return 'Sim'
            elif value_lower in ['false', 'não', 'nao', 'no', '0', 'falso']:
                return 'Não'
            else:
                return 'Não especificado'
                
        return 'Não especificado'
        
    def _format_column_value(self, value: Any, column_type: str) -> Any:
        """
        Formata valor de coluna baseado no tipo.
        
        Args:
            value: Valor a formatar
            column_type: Tipo da coluna
            
        Returns:
            Valor formatado
        """
        if value is None or value == '':
            return self._get_default_value(column_type)
            
        try:
            if column_type == 'boolean':
                if isinstance(value, bool):
                    return 'Sim' if value else 'Não'
                elif isinstance(value, str):
                    return 'Sim' if value.lower() in ['true', 'sim', 'yes', '1'] else 'Não'
                else:
                    return 'Não'
                    
            elif column_type == 'number':
                if isinstance(value, (int, float)):
                    return value
                elif isinstance(value, str):
                    try:
                        return float(value) if '.' in value else int(value)
                    except ValueError:
                        return 0
                else:
                    return 0
                    
            elif column_type == 'date':
                if isinstance(value, (date, datetime)):
                    return value.strftime('%Y-%m-%d')
                elif isinstance(value, str):
                    # Tenta parsear data
                    try:
                        parsed_date = pd.to_datetime(value)
                        return parsed_date.strftime('%Y-%m-%d')
                    except:
                        return value
                else:
                    return str(value)
                    
            else:  # text
                return str(value)
                
        except Exception:
            return self._get_default_value(column_type)
            
    def _get_default_value(self, column_type: str) -> Any:
        """
        Retorna valor padrão para tipo de coluna.
        
        Args:
            column_type: Tipo da coluna
            
        Returns:
            Valor padrão
        """
        defaults = {
            'text': 'Não especificado',
            'number': 0,
            'boolean': 'Não',
            'date': ''
        }
        return defaults.get(column_type, 'Não especificado')
        
    def _build_metadata_dataframe(self, 
                                 search_query: str,
                                 filter_criteria: str,
                                 total_articles: int,
                                 columns: List[ColumnConfig]) -> pd.DataFrame:
        """
        Constrói DataFrame com metadados da busca.
        
        Args:
            search_query: Query de busca
            filter_criteria: Critérios de filtro
            total_articles: Número total de artigos
            columns: Configurações de colunas
            
        Returns:
            DataFrame com metadados
        """
        metadata_rows = [
            {'Campo': 'Data da Análise', 'Valor': datetime.now().strftime('%Y-%m-%d %H:%M:%S')},
            {'Campo': 'Query de Busca', 'Valor': search_query},
            {'Campo': 'Critérios de Filtro', 'Valor': filter_criteria},
            {'Campo': 'Total de Artigos', 'Valor': total_articles},
            {'Campo': 'Colunas Configuradas', 'Valor': len(columns)},
            {'Campo': '', 'Valor': ''},  # Linha em branco
            {'Campo': 'CONFIGURAÇÃO DE COLUNAS', 'Valor': ''},
        ]
        
        # Adiciona detalhes das colunas
        for i, col in enumerate(columns, 1):
            metadata_rows.extend([
                {'Campo': f'Coluna {i} - Nome', 'Valor': col.name},
                {'Campo': f'Coluna {i} - Tipo', 'Valor': col.column_type},
                {'Campo': f'Coluna {i} - Descrição', 'Valor': col.description or 'Sem descrição'},
                {'Campo': '', 'Valor': ''},  # Linha em branco
            ])
            
        return pd.DataFrame(metadata_rows)
        
    def _format_workbook(self, 
                        workbook: openpyxl.Workbook,
                        df: pd.DataFrame,
                        columns: List[ColumnConfig]) -> None:
        """
        Aplica formatação avançada ao workbook.
        
        Args:
            workbook: Workbook do openpyxl
            df: DataFrame com dados
            columns: Configurações de colunas
        """
        try:
            # Formata aba principal
            ws_articles = workbook['Artigos Analisados']
            self._format_articles_sheet(ws_articles, df, columns)
            
            # Formata aba de metadados
            ws_metadata = workbook['Metadados']
            self._format_metadata_sheet(ws_metadata)
            
        except Exception as e:
            logger.warning(f"Erro na formatação do Excel: {e}")
            
    def _format_articles_sheet(self, 
                              worksheet,
                              df: pd.DataFrame,
                              columns: List[ColumnConfig]) -> None:
        """
        Formata aba de artigos.
        
        Args:
            worksheet: Worksheet do openpyxl
            df: DataFrame com dados
            columns: Configurações de colunas
        """
        # Formata cabeçalho
        for col_num, header in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
            
        # Formata dados
        for row_num in range(2, len(df) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = self.border
                
                # Alinhamento específico por tipo
                if col_num <= len(df.columns):
                    header = df.columns[col_num - 1]
                    if header in ['Citações', 'Confiança da IA']:
                        cell.alignment = Alignment(horizontal="right")
                    elif header == 'Atende Critérios':
                        cell.alignment = Alignment(horizontal="center")
                        # Cor de fundo para Sim/Não
                        if cell.value == 'Sim':
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif cell.value == 'Não':
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            
        # Ajusta largura das colunas
        column_widths = {
            'Título': 40,
            'URL': 25,
            'Resumo': 50,
            'Autores': 30,
            'Justificativa da IA': 40,
        }
        
        for col_num, header in enumerate(df.columns, 1):
            width = column_widths.get(header, 15)
            worksheet.column_dimensions[worksheet.cell(row=1, column=col_num).column_letter].width = width
            
        # Congela primeira linha
        worksheet.freeze_panes = 'A2'
        
    def _format_metadata_sheet(self, worksheet) -> None:
        """
        Formata aba de metadados.
        
        Args:
            worksheet: Worksheet do openpyxl
        """
        # Formata cabeçalho
        for col_num in [1, 2]:
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
            
        # Formata dados
        max_row = worksheet.max_row
        for row_num in range(2, max_row + 1):
            for col_num in [1, 2]:
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = self.border
                
                # Destaca seções
                if col_num == 1 and cell.value and 'CONFIGURAÇÃO' in str(cell.value).upper():
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    
        # Ajusta largura das colunas
        worksheet.column_dimensions['A'].width = 25
        worksheet.column_dimensions['B'].width = 50
    
    def _clean_dataframe_for_csv(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Limpa o DataFrame removendo caracteres que podem quebrar o CSV.
        
        Args:
            df: DataFrame original
            
        Returns:
            DataFrame limpo
        """
        cleaned_df = df.copy()
        
        # Remove ou substitui ponto e vírgula por vírgula em todas as colunas de texto
        for column in cleaned_df.columns:
            if cleaned_df[column].dtype == 'object':  # Colunas de texto
                cleaned_df[column] = cleaned_df[column].astype(str).str.replace(';', ',', regex=False)
                # Remove quebras de linha que podem quebrar o CSV
                cleaned_df[column] = cleaned_df[column].str.replace('\n', ' ', regex=False)
                cleaned_df[column] = cleaned_df[column].str.replace('\r', ' ', regex=False)
                # Remove aspas duplas extras
                cleaned_df[column] = cleaned_df[column].str.replace('"', "'", regex=False)
        
        return cleaned_df
        
    def cleanup_temp_files(self, file_path: str) -> None:
        """
        Remove arquivo temporário.
        
        Args:
            file_path: Caminho do arquivo a remover
        """
        try:
            if os.path.exists(file_path):
                os.unlink(file_path)
                logger.debug(f"Arquivo temporário removido: {file_path}")
        except Exception as e:
            logger.warning(f"Erro ao remover arquivo temporário {file_path}: {e}")